#!/usr/bin/env python3
"""
Kaset Kapasiteleri.xlsx dosyasından ATM kapasite verilerini okur
ve atm_master.json dosyasına all_in_capacity alanını ekler.
"""

import json
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl bulunamadı, yükleniyor...")
    os.system("pip install openpyxl")
    import openpyxl

# Yollar
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "ai_engine", "Kaset Kapasiteleri.xlsx")
JSON_PATH  = os.path.join(os.path.dirname(__file__), "src", "data", "atm_master.json")

# ── 1. Excel'i oku ──────────────────────────────────────────────────────────────
print(f"Excel okunuyor: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Çalışma sayfalarını listele
print(f"Sayfalar: {wb.sheetnames}")
ws = wb.active  # ilk/aktif sayfa kullan

# Başlık satırını bul ve sütun indekslerini tespit et
headers = None
capacity_map = {}   # { "FA006": 2000, ... }
row_count = 0

for row in ws.iter_rows(values_only=True):
    if headers is None:
        # Başlık satırını bul (ATM ID ve kapasite sütunlarını içermeli)
        headers = [str(c).strip().upper() if c is not None else "" for c in row]
        print(f"Başlıklar: {headers}")

        # ATM ID sütununu bul
        atm_id_col = None
        capacity_col = None
        for i, h in enumerate(headers):
            if "ATM" in h and ("ID" in h or "KOD" in h or "NO" in h):
                atm_id_col = i
            elif "ALL" in h and "IN" in h:
                capacity_col = i
            elif "KAPASİTE" in h or "KAPASITE" in h or "CAPACITY" in h:
                capacity_col = i
            elif "MODEL" in h:
                # model sütunu da kapasite için yedek olabilir
                pass

        # Otomatik tespitte başarısız olursa tüm başlıkları göster
        if atm_id_col is None or capacity_col is None:
            print("⚠️  Sütun otomatik tespit edilemedi. Mevcut başlıklar:")
            for i, h in enumerate(headers):
                print(f"  [{i}] {h!r}")
            # İlk 2 sütunu dene
            if atm_id_col is None:
                atm_id_col = 0
            if capacity_col is None:
                capacity_col = len(headers) - 1  # son sütun
            print(f"Tahmin: ATM ID sütunu=[{atm_id_col}], Kapasite sütunu=[{capacity_col}]")
        else:
            print(f"✅ ATM ID sütunu=[{atm_id_col}] '{headers[atm_id_col]}', "
                  f"Kapasite sütunu=[{capacity_col}] '{headers[capacity_col]}'")
        continue

    atm_id = row[atm_id_col]
    capacity = row[capacity_col]

    if atm_id is None:
        continue

    atm_id = str(atm_id).strip().upper()
    if not atm_id:
        continue

    try:
        cap_val = int(float(str(capacity))) if capacity is not None else None
    except (ValueError, TypeError):
        cap_val = None

    capacity_map[atm_id] = cap_val
    row_count += 1

wb.close()
print(f"Excel'den {row_count} satır okundu, {len(capacity_map)} benzersiz ATM tespit edildi.")

# ── 2. JSON'u oku ───────────────────────────────────────────────────────────────
print(f"\nJSON okunuyor: {JSON_PATH}")
with open(JSON_PATH, "r", encoding="utf-8") as f:
    atms = json.load(f)

print(f"JSON'da {len(atms)} ATM kaydı var.")

# ── 3. Kapasite alanını ekle/güncelle ───────────────────────────────────────────
matched = 0
not_found = 0
no_cap = 0

for atm in atms:
    atm_id = str(atm.get("atm_id", "")).strip().upper()
    if atm_id in capacity_map:
        cap = capacity_map[atm_id]
        atm["all_in_capacity"] = cap
        if cap is not None:
            matched += 1
        else:
            no_cap += 1
    else:
        atm["all_in_capacity"] = None
        not_found += 1

print(f"\nSonuç:")
print(f"  ✅ Kapasite eklendi  : {matched}")
print(f"  ⚠️  Excel'de bulunamadı: {not_found}")
print(f"  ❓ Kapasite değeri yok: {no_cap}")

# ── 4. JSON'u kaydet ────────────────────────────────────────────────────────────
with open(JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(atms, f, ensure_ascii=False, indent=2)

print(f"\n✅ {JSON_PATH} başarıyla güncellendi!")

# İlk 3 kaydı önizle
print("\nÖnizleme (ilk 3 kayıt):")
for atm in atms[:3]:
    print(f"  {atm['atm_id']} → all_in_capacity: {atm.get('all_in_capacity')}")
